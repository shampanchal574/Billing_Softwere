from openpyxl import Workbook, load_workbook
from datetime import datetime
import os


BASE_DIR = os.path.join(os.getcwd(), "sales_excel")
os.makedirs(BASE_DIR, exist_ok=True)


def save_sale_to_excel(customer, phone, items, total):
    """
    items = list of dicts:
    { name, price, qty }
    """

    now = datetime.now()
    filename = f"sales_{now.year}_{now.month:02d}.xlsx"
    path = os.path.join(BASE_DIR, filename)

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Date",
            "Time",
            "Customer Name",
            "Mobile No",
            "Products",
            "Total Amount"
        ])

    product_text = ", ".join(
        f"{i['name']} x{i['qty']} @ ₹{i['price']}"
        for i in items
    )

    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        customer,
        phone,
        product_text,
        total
    ])

    wb.save(path)